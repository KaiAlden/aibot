from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from app.ingestion.constitutions import extract_constitutions
from app.ingestion.forms import normalize_form


@dataclass(frozen=True)
class ProductImportRow:
    name: str
    category: str | None = None
    form: str | None = None
    description: str | None = None
    price: Decimal | None = None
    weight: str | None = None
    image_url: str | None = None
    constitutions: list[str] = field(default_factory=list)
    unknown_constitutions: list[str] = field(default_factory=list)
    ingredients: list[str] = field(default_factory=list)
    is_universal: bool = False


@dataclass(frozen=True)
class ProductImportIssue:
    sheet_name: str | None
    row_number: int | None
    field: str
    message: str
    value: str | None = None


@dataclass(frozen=True)
class ProductImportValidationResult:
    rows: list[ProductImportRow]
    issues: list[ProductImportIssue]
    total_rows: int


def parse_list_cell(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    raw = str(value).strip()
    if not raw:
        return []
    for separator in ["，", ",", "、", ";", "；", "|"]:
        raw = raw.replace(separator, ",")
    return [item.strip() for item in raw.split(",") if item.strip()]


def parse_price(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace("￥", "").replace("元", "").strip())
    except InvalidOperation:
        return None


def get_mapped_value(row: dict[str, Any], source: str | list[str]) -> Any:
    if isinstance(source, str):
        return row.get(source)
    for candidate in source:
        value = row.get(candidate)
        if value not in (None, ""):
            return value
    return None


def map_product_row(row: dict[str, Any], field_mapping: dict[str, str | list[str]]) -> ProductImportRow:
    mapped = {target: get_mapped_value(row, source) for target, source in field_mapping.items()}
    name = str(mapped.get("name") or "").strip()
    if not name:
        raise ValueError("product name is required")

    constitutions, unknown_constitutions = extract_constitutions(mapped.get("constitutions"))
    is_universal = str(mapped.get("is_universal") or "").strip() in {"1", "true", "是", "全适用"}

    return ProductImportRow(
        name=name,
        category=str(mapped["category"]).strip() if mapped.get("category") else row.get("__sheet_name"),
        form=normalize_form(str(mapped["form"])) if mapped.get("form") else None,
        description=str(mapped["description"]).strip() if mapped.get("description") else None,
        price=parse_price(mapped.get("price")),
        weight=str(mapped["weight"]).strip() if mapped.get("weight") else None,
        image_url=str(mapped["image_url"]).strip() if mapped.get("image_url") else None,
        constitutions=constitutions,
        unknown_constitutions=unknown_constitutions,
        ingredients=parse_list_cell(mapped.get("ingredients")),
        is_universal=is_universal or not constitutions,
    )


def validate_product_row(
    row: dict[str, Any],
    field_mapping: dict[str, str | list[str]],
) -> tuple[ProductImportRow | None, list[ProductImportIssue]]:
    issues: list[ProductImportIssue] = []
    try:
        product = map_product_row(row, field_mapping)
    except ValueError as exc:
        issues.append(_row_issue(row, "name", str(exc), None))
        return None, issues

    mapped_price = get_mapped_value(row, field_mapping["price"]) if "price" in field_mapping else None
    if mapped_price not in (None, "") and product.price is None:
        issues.append(_row_issue(row, "price", "invalid price", str(mapped_price)))
    for value in product.unknown_constitutions:
        issues.append(_row_issue(row, "constitutions", "unknown constitution", value))
    return product, issues


def _row_issue(row: dict[str, Any], field: str, message: str, value: str | None) -> ProductImportIssue:
    return ProductImportIssue(
        sheet_name=str(row["__sheet_name"]) if row.get("__sheet_name") else None,
        row_number=int(row["__row_number"]) if row.get("__row_number") else None,
        field=field,
        message=message,
        value=value,
    )


def read_sheet_rows(sheet: Worksheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        item["__sheet_name"] = sheet.title
        item["__row_number"] = row_number
        data_rows.append(item)
    return data_rows


def read_xlsx_rows(path: Path, all_sheets: bool = False) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = workbook.worksheets if all_sheets else [workbook.active]
        rows: list[dict[str, Any]] = []
        for sheet in sheets:
            rows.extend(read_sheet_rows(sheet))
        return rows
    finally:
        workbook.close()


def import_xlsx_products(
    path: Path,
    field_mapping: dict[str, str | list[str]],
    all_sheets: bool = False,
) -> list[ProductImportRow]:
    result = validate_xlsx_products(path, field_mapping, all_sheets=all_sheets)
    blocking_issues = [issue for issue in result.issues if is_blocking_import_issue(issue)]
    if blocking_issues:
        first_issue = blocking_issues[0]
        location = f"{first_issue.sheet_name or '-'}:{first_issue.row_number or '-'}"
        raise ValueError(f"{location} {first_issue.field}: {first_issue.message}")
    return result.rows


def is_blocking_import_issue(issue: ProductImportIssue) -> bool:
    return not (issue.field == "constitutions" and issue.message == "unknown constitution")


def validate_xlsx_products(
    path: Path,
    field_mapping: dict[str, str | list[str]],
    all_sheets: bool = False,
) -> ProductImportValidationResult:
    raw_rows = read_xlsx_rows(path, all_sheets=all_sheets)
    products: list[ProductImportRow] = []
    issues: list[ProductImportIssue] = []
    for raw_row in raw_rows:
        product, row_issues = validate_product_row(raw_row, field_mapping)
        if product is not None:
            products.append(product)
        issues.extend(row_issues)
    return ProductImportValidationResult(rows=products, issues=issues, total_rows=len(raw_rows))
